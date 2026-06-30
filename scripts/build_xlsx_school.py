# -*- coding: utf-8 -*-
"""
build_xlsx_school.py — 学校格式达成度计算表生成器
> 操作时间: 2026-06-30 15:00
> 操作agent: default

参考：丁向朝老师《大学计算机基础》达成度表格（卷三3-达成度表格.xlsx）

表格结构：
  R1: 标题（合并）
  R2: 开课专业 | 开课班级 | 开课学期
  R3: 课程名称 | 课程代码 | 任课教师
  R4: 序号|学号|学生姓名 | 毕业要求XX (D4:W4合并) | 备注
  R5: (目标分组头) 教学目标1(D5:H5) | 教学目标2(I5:M5) | 教学目标3(N5:R5) | 教学目标4(S5:W5) | 折合得分
  R6: 子项名 ×4组
  R7: 满分行
  R8+: 学生数据，折合得分 = SUM(Dn:Wn) 公式
  平均分行
  达成度行

用法：python build_xlsx_school.py config.json output.xlsx
"""
import json, openpyxl, sys, os, math
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 样式 ──
TB = Border(left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
FT   = Font(bold=True, size=14, name='宋体')
FH   = Font(bold=True, size=10, name='宋体')
FN   = Font(size=10, name='宋体')
FB   = Font(bold=True, size=10, name='宋体')
FILL_HEAD  = 'D9E2F3'
FILL_AVG   = 'F2F2F2'
FILL_FULL  = 'FFF2CC'

def mf(c):
    return PatternFill(start_color=c, end_color=c, fill_type='solid')

def sc(ws, r, c, v, font=None, fill=None, align=None, nf=None):
    cl = ws.cell(r, c, v)
    if font: cl.font = font
    if fill: cl.fill = mf(fill)
    if align: cl.alignment = align or AC
    cl.border = TB
    if nf: cl.number_format = nf
    return cl

def ms(ws, r1, r2, c1, c2, v, font=None, fill=None, align=None):
    """合并单元格"""
    cl = ws.cell(r1, c1)
    cl.value = v
    if font: cl.font = font
    if fill: cl.fill = mf(fill)
    if align: cl.alignment = align or AC
    cl.border = TB
    if r1 != r2 or c1 != c2:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    for rr in range(r1, r2+1):
        for cc in range(c1, c2+1):
            ws.cell(rr, cc).border = TB

def set_col_widths(ws, widths):
    """widths: [(col_idx, width), (col_idx, width), ...] or list of widths from col 1"""
    if isinstance(widths, list) and len(widths) > 0 and isinstance(widths[0], tuple):
        for ci, w in widths:
            ws.column_dimensions[get_column_letter(ci)].width = w
    else:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

# ── 满分配置 ──
# 5个子项 × 4目标，每个子项在每目标中的折合满分
SUB_ITEMS = [
    {"col_label": "线上学习", "raw_full": 20,  "zhe_per_target": 2.5},
    {"col_label": "阶段测试", "raw_full": 100, "zhe_per_target": 5},
    {"col_label": "学习手册", "raw_full": 20,  "zhe_per_target": 2.5},
    {"col_label": "课堂表现", "raw_full": 20,  "zhe_per_target": 2.5},
    {"col_label": "期末考试", "raw_full": 100, "zhe_per_target": 12.5},
]
N_TARGETS = 4
N_SUBITEMS = len(SUB_ITEMS)    # 5
N_COLS = N_TARGETS * N_SUBITEMS  # 20

# 列映射：col 1=序号, 2=学号, 3=姓名, 4~23=数据, 24=折合得分, 25=录入总评, 26=备注
COL_SEQ = 1    # 序号
COL_XH  = 2    # 学号
COL_NAME = 3   # 姓名
COL_DATA_START = 4  # 数据起始列
COL_DATA_END = COL_DATA_START + N_COLS - 1  # 23
COL_ZHE = COL_DATA_END + 1  # 24 折合得分
COL_LURU = COL_ZHE + 1      # 25 录入总评
COL_NOTE = COL_LURU + 1     # 26 备注

# 目标列范围
def target_col_range(t):
    """返回目标t(0-based)的数据列范围 start_col, end_col (1-based)"""
    sc = COL_DATA_START + t * N_SUBITEMS
    return sc, sc + N_SUBITEMS - 1

def build_xlsx(config_path, output_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        cfg = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '达成度'

    major = cfg['major']
    classes = '、'.join(cfg['classes'])
    indicators = [t['indicator'] for t in cfg['targets']]
    indicator_str = '、'.join(indicators)
    teacher = cfg['teacher']
    course_name = cfg['course_name']
    course_code = cfg['course_code']
    term = cfg['term']
    students = cfg['students']
    n_stu = len(students)

    # 折合满分
    full_per_target = sum(si['zhe_per_target'] for si in SUB_ITEMS)  # 25
    full_total = full_per_target * N_TARGETS  # 100

    # ── 列宽 ──
    ws.column_dimensions['A'].width = 6    # 序号
    ws.column_dimensions['B'].width = 14   # 学号
    ws.column_dimensions['C'].width = 10   # 姓名
    for cc in range(COL_DATA_START, COL_DATA_END + 1):
        ws.column_dimensions[get_column_letter(cc)].width = 6.5
    ws.column_dimensions[get_column_letter(COL_ZHE)].width = 9  # 折合得分
    ws.column_dimensions[get_column_letter(COL_LURU)].width = 9  # 录入总评
    ws.column_dimensions[get_column_letter(COL_NOTE)].width = 8  # 备注

    # ═══ Row 1: 标题 ═══
    r = 1
    ms(ws, r, r, 1, COL_NOTE, f'《{course_name}》课程教学目标达成计算表', font=FT, fill=FILL_HEAD)
    ws.row_dimensions[r].height = 28

    # ═══ Row 2: 专业/班级/学期 ═══
    r = 2
    ms(ws, r, r, 1, 3, f'开课专业：{major}', font=FH, align=AL)
    ms(ws, r, r, COL_DATA_START, COL_DATA_START + 4, f'开课班级：{classes}', font=FH, align=AL)
    ms(ws, r, r, COL_DATA_START + 6, COL_DATA_START + 11, f'开课学期：{term}', font=FH, align=AL)

    # ═══ Row 3: 课程名/代码/教师 ═══
    r = 3
    ms(ws, r, r, 1, 3, f'课程名称：{course_name}', font=FH, align=AL)
    ms(ws, r, r, COL_DATA_START, COL_DATA_START + 4, f'课程代码：{course_code}', font=FH, align=AL)
    ms(ws, r, r, COL_DATA_START + 6, COL_DATA_START + 11, f'任课教师：{teacher}', font=FH, align=AL)

    # ═══ Row 4: 表头 - 序号/学号/姓名/毕业要求/备注 ═══
    r = 4
    sc(ws, r, COL_SEQ, '序号', font=FH, fill=FILL_HEAD)
    sc(ws, r, COL_XH, '学号', font=FH, fill=FILL_HEAD)
    sc(ws, r, COL_NAME, '学生姓名', font=FH, fill=FILL_HEAD)
    ms(ws, r, r, COL_DATA_START, COL_DATA_END, f'毕业要求{indicator_str}', font=FH, fill=FILL_HEAD)
    sc(ws, r, COL_ZHE, '折合\n得分', font=FH, fill=FILL_HEAD)
    sc(ws, r, COL_LURU, '录入\n总评', font=FH, fill=FILL_HEAD)
    sc(ws, r, COL_NOTE, '备注', font=FH, fill=FILL_HEAD)
    ws.row_dimensions[r].height = 32

    # ═══ Row 5: 教学目标分组头 ═══
    r = 5
    sc(ws, r, COL_SEQ, '', fill=FILL_HEAD)
    sc(ws, r, COL_XH, '', fill=FILL_HEAD)
    sc(ws, r, COL_NAME, '', fill=FILL_HEAD)
    for t in range(N_TARGETS):
        sc_, ec_ = target_col_range(t)
        ms(ws, r, r, sc_, ec_, f'{cfg["targets"][t]["name"]}（{indicators[t]}）',
           font=FH, fill=FILL_HEAD)
    sc(ws, r, COL_ZHE, '', fill=FILL_HEAD)
    sc(ws, r, COL_LURU, '', fill=FILL_HEAD)
    sc(ws, r, COL_NOTE, '', fill=FILL_HEAD)
    ws.row_dimensions[r].height = 28

    # ═══ Row 6: 子项表头 ═══
    r = 6
    sc(ws, r, COL_SEQ, '', fill=FILL_HEAD)
    sc(ws, r, COL_XH, '', fill=FILL_HEAD)
    sc(ws, r, COL_NAME, '', fill=FILL_HEAD)
    for t in range(N_TARGETS):
        sc_, ec_ = target_col_range(t)
        for si_idx, si in enumerate(SUB_ITEMS):
            col = sc_ + si_idx
            sc(ws, r, col, si['col_label'], font=FH, fill=FILL_HEAD)
    sc(ws, r, COL_ZHE, '折合\n得分', font=FH, fill=FILL_HEAD)
    sc(ws, r, COL_LURU, '录入\n总评', font=FH, fill=FILL_HEAD)
    sc(ws, r, COL_NOTE, '', fill=FILL_HEAD)
    ws.row_dimensions[r].height = 28

    # ═══ Row 7: 满分行 ═══
    r = 7
    sc(ws, r, COL_SEQ, '', fill=FILL_FULL)
    sc(ws, r, COL_XH, '', fill=FILL_FULL)
    sc(ws, r, COL_NAME, '', fill=FILL_FULL)
    for t in range(N_TARGETS):
        sc_, ec_ = target_col_range(t)
        for si_idx, si in enumerate(SUB_ITEMS):
            col = sc_ + si_idx
            sc(ws, r, col, si['zhe_per_target'], font=FH, fill=FILL_FULL)
    sc(ws, r, COL_ZHE, full_total, font=FH, fill=FILL_FULL)
    sc(ws, r, COL_LURU, 100, font=FH, fill=FILL_FULL)
    sc(ws, r, COL_NOTE, '', fill=FILL_FULL)
    ws.row_dimensions[r].height = 22

    # 红色字体 — 标记调整过的单元格
    RED_FONT = Font(name='微软雅黑', size=10, color='FF0000')
    RED_FONT_BOLD = Font(name='微软雅黑', size=10, color='FF0000', bold=True)

    # ═══ 学生数据行（R8 起）═══
    DATA_START_ROW = 8
    for i, stu in enumerate(students):
        r = DATA_START_ROW + i
        sc(ws, r, COL_SEQ, i + 1, font=FN)
        sc(ws, r, COL_XH, stu['xh'], font=FN, align=AL)
        sc(ws, r, COL_NAME, stu['name'], font=FN)

        raw = stu['raw_scores']
        luru_zp = stu.get('luru_zp', 0)

        # 阶段测试映射：阶段测试1→目标1, 2→目标2, ...
        st_keys = {t: f'阶段测试{t+1}' for t in range(N_TARGETS)}

        # ── 第一遍：计算所有20列折合值 ──
        zhe_vals = {}  # col -> rounded zhe_score
        for t in range(N_TARGETS):
            sc_, ec_ = target_col_range(t)
            for si_idx, si in enumerate(SUB_ITEMS):
                col = sc_ + si_idx
                label = si['col_label']
                raw_key = st_keys[t] if label == '阶段测试' else label
                raw_score = raw.get(raw_key, 0)
                zhe_score = raw_score / si['raw_full'] * si['zhe_per_target']
                zhe_vals[col] = round(zhe_score, 4)

        # ── 折合合计 vs 录入总评 ──
        zhe_sum = sum(zhe_vals.values())
        gap = luru_zp - zhe_sum

        if abs(gap) <= 2:
            # 无需调整，直接写入
            for col, val in zhe_vals.items():
                sc(ws, r, col, round(val, 4), font=FN, nf='0.000')
            adjusted = False
            note_text = ''
        else:
            # 超过2分 → 只调平时成绩（线上/测试/课堂/手册），不动期末
            remaining = gap
            adjusted_cols = set()

            # 排序：阶段测试(5)优先，然后线上/课堂/手册(2.5)，排除期末(12.5)
            cells = []
            for col, val in zhe_vals.items():
                si = SUB_ITEMS[(col - COL_DATA_START) % N_SUBITEMS]
                if si['col_label'] == '期末考试':
                    continue  # 不动期末
                cells.append((col, si, val))
            cells.sort(key=lambda x: x[1]['zhe_per_target'], reverse=True)

            for col, si, val in cells:
                if abs(remaining) < 0.001:
                    break
                if remaining > 0:
                    free = si['zhe_per_target'] - val
                    adj = min(remaining, free)
                else:
                    free = val
                    adj = max(remaining, -free)
                if abs(adj) > 0.0001:
                    zhe_vals[col] = round(val + adj, 6)
                    adjusted_cols.add(col)
                    remaining -= adj

            # 写入：改过的红色，没改的黑色
            for col, val in zhe_vals.items():
                font = RED_FONT if col in adjusted_cols else FN
                sc(ws, r, col, round(val, 4), font=font, nf='0.000')
            adjusted = True
            note_text = '已调' if abs(remaining) < 0.01 else f'差{remaining:.1f}'

        # 折合得分 = SUM
        zhe_formula = f'=SUM({get_column_letter(COL_DATA_START)}{r}:{get_column_letter(COL_DATA_END)}{r})'
        sc(ws, r, COL_ZHE, zhe_formula, font=RED_FONT_BOLD if adjusted else FB, nf='0.0')
        # 录入总评（与录入成绩.xlsx完全一致）
        sc(ws, r, COL_LURU, luru_zp, font=FN, nf='0.0')
        sc(ws, r, COL_NOTE, note_text, font=RED_FONT_BOLD if adjusted and note_text != '已调' else FN)

    # ═══ 平均分行 ═══
    r_avg = DATA_START_ROW + n_stu
    sc(ws, r_avg, COL_SEQ, '平均分', font=FB, fill=FILL_AVG)
    sc(ws, r_avg, COL_XH, '', fill=FILL_AVG)
    sc(ws, r_avg, COL_NAME, '', fill=FILL_AVG)
    for col in range(COL_DATA_START, COL_DATA_END + 1):
        col_l = get_column_letter(col)
        sc(ws, r_avg, col,
           f'=AVERAGE({col_l}{DATA_START_ROW}:{col_l}{DATA_START_ROW + n_stu - 1})',
           font=FB, fill=FILL_AVG, nf='0.000')
    sc(ws, r_avg, COL_ZHE,
       f'=AVERAGE({get_column_letter(COL_ZHE)}{DATA_START_ROW}:{get_column_letter(COL_ZHE)}{DATA_START_ROW + n_stu - 1})',
       font=FB, fill=FILL_AVG, nf='0.00')
    sc(ws, r_avg, COL_LURU,
       f'=AVERAGE({get_column_letter(COL_LURU)}{DATA_START_ROW}:{get_column_letter(COL_LURU)}{DATA_START_ROW + n_stu - 1})',
       font=FB, fill=FILL_AVG, nf='0.0')
    sc(ws, r_avg, COL_NOTE, '', fill=FILL_AVG)

    # ═══ 课程教学目标达成度 ═══
    r_dc = r_avg + 1
    sc(ws, r_dc, COL_SEQ, '课程教学目标达成度', font=FB, fill=FILL_AVG)
    sc(ws, r_dc, COL_XH, '', fill=FILL_AVG)
    sc(ws, r_dc, COL_NAME, '', fill=FILL_AVG)
    for t in range(N_TARGETS):
        sc_, ec_ = target_col_range(t)
        # SUM(avg_cols) / SUM(full_cols)
        avg_cols = ','.join([f'{get_column_letter(c)}{r_avg}' for c in range(sc_, ec_ + 1)])
        full_cols = ','.join([f'{get_column_letter(c)}7' for c in range(sc_, ec_ + 1)])
        formula = f'=SUM({avg_cols})/SUM({full_cols})'
        sc(ws, r_dc, sc_, formula, font=FB, fill=FILL_AVG, nf='0.0000')
        # fill remaining cols in this target group
        for cc in range(sc_ + 1, ec_ + 1):
            sc(ws, r_dc, cc, '', fill=FILL_AVG)
    sc(ws, r_dc, COL_ZHE, '——', font=FB, fill=FILL_AVG)
    sc(ws, r_dc, COL_LURU, '——', font=FB, fill=FILL_AVG)
    sc(ws, r_dc, COL_NOTE, '', fill=FILL_AVG)
    ws.row_dimensions[r_dc].height = 22

    # ═══ 课程达成度 ═══
    r_cdc = r_dc + 1
    sc(ws, r_cdc, COL_SEQ, '课程达成度', font=FB, fill=FILL_AVG)
    sc(ws, r_cdc, COL_XH, '', fill=FILL_AVG)
    sc(ws, r_cdc, COL_NAME, '', fill=FILL_AVG)

    # AVERAGE of 4 target达成度 cells
    dc_cells = ','.join([f'{get_column_letter(target_col_range(t)[0])}{r_dc}' for t in range(N_TARGETS)])
    sc(ws, r_cdc, COL_DATA_START, f'=AVERAGE({dc_cells})', font=FB, fill=FILL_AVG, nf='0.0000')
    for cc in range(COL_DATA_START + 1, COL_DATA_END + 1):
        sc(ws, r_cdc, cc, '', fill=FILL_AVG)
    sc(ws, r_cdc, COL_ZHE, '简单平均', font=FB, fill=FILL_AVG)
    sc(ws, r_cdc, COL_LURU, '', fill=FILL_AVG)
    sc(ws, r_cdc, COL_NOTE, '', fill=FILL_AVG)
    ws.row_dimensions[r_cdc].height = 22

    # ── 保存 ──
    ws.freeze_panes = f'A{DATA_START_ROW}'
    wb.save(output_path)

    # ── 验算 ──
    # 用 Python 验算（不依赖 Excel 打开后计算）
    verify(cfg)

    return {
        'output': os.path.abspath(output_path),
        'row_count': n_stu,
        'target_dcs': {},
        'course_dc': 0,
    }

def verify(cfg):
    """用 Python 验算达成度（不依赖 Excel 公式求值）"""
    students = cfg['students']
    n_stu = len(students)
    
    fulls = []
    for t in range(N_TARGETS):
        for si in SUB_ITEMS:
            fulls.append(si['zhe_per_target'])
    
    # 计算每个学生每列的折合分
    all_zhe = []
    for stu in students:
        raw = stu['raw_scores']
        row = []
        for t in range(N_TARGETS):
            for si in SUB_ITEMS:
                label = si['col_label']
                if label == '阶段测试':
                    raw_key = f'阶段测试{t+1}'
                else:
                    raw_key = label
                raw_score = raw.get(raw_key, 0)
                zhe_score = raw_score / si['raw_full'] * si['zhe_per_target']
                row.append(zhe_score)
        all_zhe.append(row)
    
    # 每列平均
    avgs = [sum(all_zhe[i][c] for i in range(n_stu)) / n_stu for c in range(N_COLS)]
    
    # Per-target达成度
    print(f'\n=== {cfg["major"]} 达成度验算 ({n_stu}人) ===')
    for t in range(N_TARGETS):
        sc_ = t * N_SUBITEMS
        ec_ = sc_ + N_SUBITEMS
        dc = sum(avgs[sc_:ec_]) / sum(fulls[sc_:ec_])
        print(f'  {cfg["targets"][t]["name"]}({cfg["targets"][t]["indicator"]}): {dc:.4f}')
    
    course_dc = sum(sum(avgs[sc:sc+N_SUBITEMS]) / sum(fulls[sc:sc+N_SUBITEMS]) 
                    for sc in range(0, N_COLS, N_SUBITEMS)) / N_TARGETS
    print(f'  课程达成度: {course_dc:.4f}')
    
    # 检查折合总分范围
    zhe_totals = [sum(row) for row in all_zhe]
    max_zhe = max(zhe_totals)
    min_zhe = min(zhe_totals)
    ft = sum(si['zhe_per_target'] for si in SUB_ITEMS) * N_TARGETS
    print(f'  折合总分范围: {min_zhe:.2f} ~ {max_zhe:.2f} (满分{ft})')
    if max_zhe > ft + 0.01:
        print(f'  ⚠️ 警告: 有学生折合总分超过{ft}!')


if __name__ == '__main__':
    build_xlsx(sys.argv[1], sys.argv[2])
