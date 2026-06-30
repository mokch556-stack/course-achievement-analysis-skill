#!/usr/bin/env python3
"""
build_xlsx.py v4 — 达成度计算表（14列：平时子项单独列 + 期末整体列）
操作时间: 2026-06-28 22:40  操作agent: default
"""
import sys, json, argparse
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要 openpyxl 库"); sys.exit(1)

def mf(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')
FG, FH = 'F2F2F2', 'D9E2F3'
FT = Font(bold=True, size=14)
FHDR = Font(bold=True, size=10)
FN = Font(size=10)
FB = Font(bold=True, size=10)
AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
TB = Border(left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

def sc(ws, r, c, v, font=None, fill=None, align=None, nf=None):
    cl = ws.cell(r, c, v)
    if font: cl.font = font
    if fill: cl.fill = mf(fill)
    if align: cl.alignment = align
    if nf: cl.number_format = nf
    cl.border = TB; return cl

def ms(ws, r1, r2, c1, c2, v, font=None, fill=None, align=None):
    if r1 != r2 or c1 != c2: ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cl = ws.cell(r1, c1); cl.value = v
    if font: cl.font = font
    if fill: cl.fill = mf(fill)
    if align: cl.alignment = align
    cl.border = TB
    for rr in range(r1, r2+1):
        for cc in range(c1, c2+1): ws.cell(rr, cc).border = TB

SUB_NAMES = ['阶段测试', '课堂笔记', '项目方案', '课程竞赛']
# 子项折合满分(每个目标)：阶段测试7.5, 课堂笔记2.5, 项目方案10, 课程竞赛5
ZH_MAP = {'阶段测试': 7.5, '课堂笔记': 2.5, '项目方案': 10, '课程竞赛': 5}
# 子项原始满分：阶段测试30, 课堂笔记10, 项目方案40, 课程竞赛20
ORIG_MAP = {'阶段测试': 30, '课堂笔记': 10, '项目方案': 40, '课程竞赛': 20}

def zh(target_id, sub_name):
    return ZH_MAP.get(sub_name, 0)

def get_qm_tf(qm_i, tid):
    """获取某目标的期末满分值"""
    for item in qm_i:
        for d in item.get('target_dist', []):
            if d['target'] == tid:
                return d['score']
    return 0

def build_xlsx(config_path, output_path):
    with open(config_path, 'r', encoding='utf-8') as f: cfg = json.load(f)
    wb = openpyxl.Workbook(); ws1 = wb.active; ws1.title = '达成度计算表'
    tg = cfg['targets']; nt = len(tg); tfs = cfg.get('target_full_scores', {})
    si = cfg.get('sub_items', {}); ps_i = si.get('平时', []); qm_i = si.get('期末', [])
    npt = 4 + 1  # 每个目标4平时子项+1期末
    tc = 3 + nt * npt + 1  # 总列数

    # R1: 标题
    ms(ws1, 1, 1, 1, tc, f"《{cfg['course_name']}》课程教学目标达成计算表", FT, align=AC)
    cs = '、'.join(cfg.get('classes', []))
    mid = tc // 2

    # R2: 课程信息
    ms(ws1, 2, 2, 1, mid,
       f"开课专业：{cfg.get('major','电子商务')}    开课班级：{cs}    开课学期：{cfg.get('term','')}", FN, align=AL)
    ms(ws1, 2, 2, mid+1, tc,
       f"课程性质：选修课    学时学分：{cfg.get('hours')}学时，{cfg.get('credits')}学分", FN, align=AL)

    # R3: 考核构成
    ppt = 16 // nt
    ps_full = sum(i['full_score'] for i in ps_i)
    qm_full = sum(i['full_score'] for i in qm_i)
    tp = []
    for ti, t in enumerate(tg):
        tf = tfs.get(f"目标{t['id']}", 0)
        tp.append(f"目标{t['id']}=课点{ti*ppt+1}-{(ti+1)*ppt}({round(tf)}%)")
    ms(ws1, 3, 3, 1, mid,
       f"课程名称：{cfg['course_name']}    课程代码：{cfg.get('course_code','')}    教师：{cfg.get('teacher','')}", FN, align=AL)
    ms(ws1, 3, 3, mid+1, tc,
       f"考核构成：平时{int(ps_full)}%+期末{int(qm_full)}%    课点归属：{'，'.join(tp)}", FN, align=AL)

    # R4: 满分分解
    zp = []
    for t in tg:
        tid = t['id']; tf = tfs.get(f"目标{tid}", 0)
        td = '+'.join([f"{n}{zh(tid,n)}" for n in SUB_NAMES])
        qmt = get_qm_tf(qm_i, tid)
        zp.append(f"目标{tid}={int(tf)}（{td}+期末{int(qmt)}）")
    pd = '+'.join([f"{i['name']}{int(i['full_score'])}" for i in ps_i])
    qd = '+'.join([f"{i['name']}{int(i['full_score'])}" for i in qm_i])
    ms(ws1, 4, 4, 1, tc,
       f"折合满分：{'，'.join(zp)}  |  平时原始满分{int(ps_full)}={pd}，期末原始满分{int(qm_full)}={qd}", FN, align=AL)

    # R5: 空行
    for c in range(1, tc+1): ws1.cell(5, c).border = TB

    # R6: 毕业要求
    ms(ws1, 6, 6, 1, 3, '毕业要求', FHDR, FH, AC)
    col = 4
    for ti, t in enumerate(tg):
        ec = col + npt - 1
        ms(ws1, 6, 6, col, ec, f"毕业要求{ti+1}", FHDR, FH, AC); col = ec + 1
    sc(ws1, 6, tc, '折合总分', FHDR, FH, AC)

    # R7: 教学目标
    ms(ws1, 7, 7, 1, 3, '学生信息', FHDR, FH, AC)
    col = 4
    for ti, t in enumerate(tg):
        ec = col + npt - 1; tf = tfs.get(f"目标{t['id']}", 50)
        ms(ws1, 7, 7, col, ec, f"教学目标{ti+1}（课点{ti*ppt+1}-{(ti+1)*ppt}，{round(tf)}%）", FHDR, FH, AC)
        col = ec + 1
    sc(ws1, 7, tc, '课程', FHDR, FH, AC)

    # R8: 考核项名
    sc(ws1, 8, 1, '序号', FHDR, FH, AC); sc(ws1, 8, 2, '学号', FHDR, FH, AC); sc(ws1, 8, 3, '姓名', FHDR, FH, AC)
    col = 4
    for t in tg:
        tid = t['id']
        for n in SUB_NAMES:
            sc(ws1, 8, col, f"{n}（{zh(tid, n)}分）", FHDR, FH, AC); col += 1
        qmt = get_qm_tf(qm_i, tid)
        sc(ws1, 8, col, f"期末项目作业（{int(qmt)}分）", FHDR, FH, AC); col += 1
    sc(ws1, 8, tc, '折合总分（100分）', FHDR, FH, AC)

    # R9: 满分值
    for c in range(1, 4): ws1.cell(9, c).border = TB
    col = 4
    for t in tg:
        tid = t['id']
        for n in SUB_NAMES: sc(ws1, 9, col, zh(tid, n), FB, align=AC); col += 1
        qmt = get_qm_tf(qm_i, tid)
        sc(ws1, 9, col, int(qmt), FB, align=AC); col += 1
    sc(ws1, 9, tc, 100, FB, align=AC)

    # R10+: 学生数据（R9满分值后直接数据，无冗余"不分子项"行）
    students = cfg.get('students', [])
    if not students: wb.save(output_path); return
    ds = 10; de = 9 + len(students)
    for si, stu in enumerate(students):
        r = ds + si
        sc(ws1, r, 1, stu.get('id', si+1), FN, align=AC)
        sc(ws1, r, 2, str(stu.get('xuehao','')), FN, align=AC)
        sc(ws1, r, 3, stu.get('name',''), FN, align=AC)
        col = 4; ss = stu.get('sub_scores', {})
        for t in tg:
            tid = t['id']
            for n in SUB_NAMES:
                sc(ws1, r, col, ss.get(f'{n}-目标{tid}', 0), FN, align=AC, nf='0.00'); col += 1
            sc(ws1, r, col, stu['scores'].get(f'期末项目-目标{tid}', 0), FN, align=AC, nf='0.00'); col += 1
        sc(ws1, r, tc, round(stu.get('zonghe',0), 2), FB, align=AC, nf='0.00')

    # 平均行
    ra = de + 1
    sc(ws1, ra, 1, '', FB, FG, AC); sc(ws1, ra, 2, '', FB, FG, AC)
    ms(ws1, ra, ra, 3, 3, '平均分', FB, FG, AC)
    for c in range(4, tc+1):
        cl = get_column_letter(c)
        sc(ws1, ra, c, f"=IFERROR(AVERAGE({cl}{ds}:{cl}{de}),0)", FB, FG, AC, nf='0.00')

    # 达成度行
    rach = ra + 1
    ms(ws1, rach, rach, 1, 3, '课程教学目标达成度', FB, FG, AC)
    col = 4
    for t in tg:
        tid = t['id']; tf = tfs.get(f"目标{tid}", 0)
        ec = col + npt - 1
        # 该目标所有子列平均值之和 / 目标满分
        cl_parts = [f"AVERAGE({get_column_letter(col+co)}{ds}:{get_column_letter(col+co)}{de})" for co in range(npt)]
        formula = f"=IFERROR(({' + '.join(cl_parts)})/{tf},0)"
        ms(ws1, rach, rach, col, ec, '', FB, FG, AC)
        sc(ws1, rach, col, formula, FB, FG, AC, nf='0.00')
        col = ec + 1
    # 课程达成度 = 各目标达成度的加权平均（按满分加权）
    g1_col = get_column_letter(4)
    g2_col = get_column_letter(4 + npt)
    tf1 = tfs.get("目标1", 0)
    tf2 = tfs.get("目标2", 0)
    tt = tf1 + tf2
    sc(ws1, rach, tc, f"=IFERROR(({g1_col}{rach}*{tf1}+{g2_col}{rach}*{tf2})/{tt},0)", FB, FG, AC, nf='0.00')

    # 实际计算达成度值（用于显示在总结行）
    def calc_target_chengdadu(students, tg, tid, tfs):
        """计算某目标的达成度"""
        tf = tfs.get(f"目标{tid}", 0)
        total_avg = 0
        for stu in students:
            s = 0
            for n in SUB_NAMES:
                s += stu.get('sub_scores',{}).get(f'{n}-目标{tid}',0)
            s += stu['scores'].get(f'期末项目-目标{tid}',0)
            total_avg += s
        n = len(students)
        return round(total_avg / n / tf, 2) if n and tf else 0

    def calc_course_chengdadu(students, tg, tfs):
        """计算课程达成度 = 各目标达成度的加权平均"""
        t1_cd = calc_target_chengdadu(students, tg, 1, tfs)
        t2_cd = calc_target_chengdadu(students, tg, 2, tfs)
        tf1 = tfs.get("目标1", 0)
        tf2 = tfs.get("目标2", 0)
        tt = tf1 + tf2
        return round((t1_cd * tf1 + t2_cd * tf2) / tt, 2) if tt else 0

    # 总结行（显示实际数值）
    rs = rach + 1
    t1_cd = calc_target_chengdadu(students, tg, 1, tfs)
    t2_cd = calc_target_chengdadu(students, tg, 2, tfs)
    course_cd = calc_course_chengdadu(students, tg, tfs)
    ms(ws1, rs, rs, 1, tc,
       f"课程达成度：{course_cd}（目标1={t1_cd}，目标2={t2_cd}）", FB, FG, AC)

    # 分班行
    classes = cfg.get('classes', [])
    for ci, cls_name in enumerate(classes):
        rc = rs + 1 + ci
        cls_stu = [s for s in students if s.get('class','') == cls_name]
        if cls_stu:
            cp = []; col = 4
            for t in tg:
                tid = t['id']; tf = tfs.get(f"目标{tid}", 0)
                total = 0
                for stu in cls_stu:
                    for n in SUB_NAMES:
                        total += stu.get('sub_scores',{}).get(f'{n}-目标{tid}',0)
                    total += stu['scores'].get(f'期末项目-目标{tid}',0)
                a = round(total/(len(cls_stu)*tf), 2) if cls_stu else 0
                cp.append(f"目标{tid}={a}"); col += npt
            ms(ws1, rc, rc, 1, tc, f"{cls_name}达成度：{'，'.join(cp)}", FB, FG, AC)
        else:
            ms(ws1, rc, rc, 1, tc, f"{cls_name}达成度：目标1=-，目标2=-", FB, FG, AC)

    # 列宽
    ws1.column_dimensions['A'].width = 6; ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 10
    for c in range(4, tc+1): ws1.column_dimensions[get_column_letter(c)].width = 15

    # ===== Sheet2: 满分来源详解 =====
    ws2 = wb.create_sheet('满分来源详解')
    for ci, h in enumerate(['考核项目', '原始满分', '目标1\n（课点50%）', '目标2\n（课点50%）', '说明']):
        sc(ws2, 1, ci+1, h, FHDR, FH, AC)
    r = 2
    for item in ps_i:
        t1 = next((d['score'] for d in item.get('target_dist',[]) if d['target']==1), '')
        t2 = next((d['score'] for d in item.get('target_dist',[]) if d['target']==2), '')
        sc(ws2, r, 1, f"{item['name']}（平时）", FN, align=AC)
        sc(ws2, r, 2, int(item['full_score']), FN, align=AC)
        sc(ws2, r, 3, t1, FN, align=AC)
        sc(ws2, r, 4, t2, FN, align=AC)
        sc(ws2, r, 5, "课点归属各50%" if t1 and t2 else (f"归属于目标1" if t1 else "归属于目标2"), FN, align=AC)
        r += 1
    pst1 = int(sum(i['full_score'] for i in ps_i)/2)
    ms(ws2, r, r, 1, 1, '【小计】平时成绩', FB, FG, AC)
    sc(ws2, r, 2, int(ps_full), FB, FG, AC)
    sc(ws2, r, 3, pst1, FB, FG, AC)
    sc(ws2, r, 4, pst1, FB, FG, AC)
    sc(ws2, r, 5, f"平时原始满分{int(ps_full)}×各目标50%", FB, FG, AC)
    r += 1

    # 期末项目
    for item in qm_i:
        t1 = next((d['score'] for d in item.get('target_dist',[]) if d['target']==1), '')
        t2 = next((d['score'] for d in item.get('target_dist',[]) if d['target']==2), '')
        sc(ws2, r, 1, f"{item['name']}（期末）", FN, align=AC)
        sc(ws2, r, 2, int(item['full_score']), FN, align=AC)
        sc(ws2, r, 3, t1, FN, align=AC)
        sc(ws2, r, 4, t2, FN, align=AC)
        sc(ws2, r, 5, "课点归属各50%" if t1 and t2 else (f"归属于目标1" if t1 else "归属于目标2"), FN, align=AC)
        r += 1
    qmht = int(sum(i['full_score'] for i in qm_i)/2)
    ms(ws2, r, r, 1, 1, '【小计】期末成绩', FB, FG, AC)
    sc(ws2, r, 2, int(qm_full), FB, FG, AC)
    sc(ws2, r, 3, qmht, FB, FG, AC)
    sc(ws2, r, 4, qmht, FB, FG, AC)
    sc(ws2, r, 5, f"期末原始满分{int(qm_full)}×各目标50%", FB, FG, AC)
    r += 1

    ms(ws2, r, r, 1, 1, '【合计】满分值', FB)
    sc(ws2, r, 2, int(ps_full+qm_full), FB, align=AC)
    tf1 = tfs.get("目标1", int(pst1+qmht))
    tf2 = tfs.get("目标2", int(pst1+qmht))
    sc(ws2, r, 3, tf1, FB, align=AC)
    sc(ws2, r, 4, tf2, FB, align=AC)
    sc(ws2, r, 5, f"目标1={tf1}分、目标2={tf2}分（来自教学大纲考核表）", FB, align=AC)

    # 列宽
    ws2.column_dimensions['A'].width = 22; ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12; ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 30

    wb.save(output_path)
    print(f"[OK] 达成度计算表已生成: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='build_xlsx v4 — 14列子项布局')
    parser.add_argument('config', help='配置文件路径')
    parser.add_argument('-o', '--output', default=None, help='输出xlsx路径')
    args = parser.parse_args()
    if args.output is None:
        args.output = args.config.replace('_config_', '达成度计算表_').replace('.json', '.xlsx')
    build_xlsx(args.config, args.output)


if __name__ == '__main__':
    main()
