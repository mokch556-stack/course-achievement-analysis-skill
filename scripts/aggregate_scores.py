"""
aggregate_scores.py — 从课程档案目录归集所有子项成绩（v2-修复版）

操作时间: 2026-06-29
操作agent: default

归集策略：
- 互金251：优先用"平时成绩报告单"（含4子项折合分）
- 工管251：从各子项源文件逐一提取
- 期末成绩：从PDF成绩登记表
"""
import openpyxl, pdfplumber, json, os, re
from collections import defaultdict

def normalize_xh(xh):
    return re.sub(r'\D', '', str(xh or ''))

def read_pscj(path):
    """读平时成绩报告单 → {xh: {线上学习, 课堂表现, 阶段测试, 学习手册, 班级}}
    
    兼容三页：模版(互金251), Sheet1(工管251), Sheet2(总评,跳过)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    scores = {}
    
    for sname in ['模版', 'Sheet1']:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        for r in range(5, ws.max_row + 1):
            xh = normalize_xh(ws.cell(r, 2).value)
            if not xh or len(xh) < 10:
                continue
            name = str(ws.cell(r, 3).value or '').strip()
            if not name:
                continue
            c4 = ws.cell(r, 4).value
            c5 = ws.cell(r, 5).value
            c6 = ws.cell(r, 6).value
            c7 = ws.cell(r, 7).value
            scores[xh] = {
                'name': name,
                'raw_scores': {
                    '线上学习': float(c4 if c4 is not None else 0),
                    '课堂表现': float(c5 if c5 is not None else 0),
                    '阶段测试': float(c6 if c6 is not None else 0),
                    '学习手册': float(c7 if c7 is not None else 0),
                },
                '_sheet': sname,
            }
    return scores

def read_shouce(path):
    """读手册成绩 → {xh: 学习手册(汇总列, 20分满分)}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    scores = {}
    for r in range(4, ws.max_row + 1):
        xh = normalize_xh(ws.cell(r, 2).value)
        if not xh or len(xh) < 10:
            continue
        v = ws.cell(r, 7).value  # 汇总(20%)
        if v is not None:
            scores[xh] = round(float(str(v).replace(',', '')), 2)
    return scores

def read_ketang(path):
    """读课堂表现成绩 → {xh: 课堂表现(汇总列, 20分满分)}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    scores = {}
    for r in range(4, ws.max_row + 1):
        xh = normalize_xh(ws.cell(r, 2).value)
        if not xh or len(xh) < 10:
            continue
        v = ws.cell(r, 9).value  # 汇总
        if v is not None:
            scores[xh] = round(float(str(v).replace(',', '')), 2)
    return scores

def read_stage_tests(stage_dir):
    """读阶段测试1-4 → {xh: {阶段测试1: raw1, 阶段测试2: raw2, 阶段测试3: raw3, 阶段测试4: raw4}}"""
    scores = defaultdict(list)
    for fname in sorted(os.listdir(stage_dir)):
        if not fname.endswith('.xlsx'):
            continue
        path = os.path.join(stage_dir, fname)
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        for r in range(4, ws.max_row + 1):
            xh = normalize_xh(ws.cell(r, 2).value)
            if not xh or len(xh) < 10:
                continue
            v = ws.cell(r, 9).value
            if v and v != '未参加考试' and str(v).strip():
                try:
                    scores[xh].append(float(str(v)))
                except (ValueError, TypeError):
                    scores[xh].append(0)
            else:
                scores[xh].append(0)
    
    result = {}
    for xh, tests in scores.items():
        if len(tests) >= 4:
            result[xh] = {
                '阶段测试1': round(float(tests[0]), 1),
                '阶段测试2': round(float(tests[1]), 1),
                '阶段测试3': round(float(tests[2]), 1),
                '阶段测试4': round(float(tests[3]), 1),
            }
        elif tests:
            # 不足4个，补齐0
            padded = tests + [0] * (4 - len(tests))
            result[xh] = {
                '阶段测试1': round(float(padded[0]), 1),
                '阶段测试2': round(float(padded[1]), 1),
                '阶段测试3': round(float(padded[2]), 1),
                '阶段测试4': round(float(padded[3]), 1),
            }
    return result

def read_online(path):
    """读线上成绩报告单 → {xh: 20分满分}
    
    结构：R5的C4/C6/C8/...是视频名，数据行的同名列为"已完成"/空。
    相邻列(C5/C7/...)也存状态，必须只数视频名列避免重复。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    # 找到所有视频名列的索引
    video_cols = []
    for c in range(4, ws.max_column + 1):
        v = ws.cell(5, c).value
        if v and ('.mp4' in str(v) or '.flv' in str(v)):
            video_cols.append(c)
    
    total_videos = len(video_cols)
    scores = {}
    for r in range(6, ws.max_row + 1):
        xh = normalize_xh(ws.cell(r, 2).value)
        if not xh or len(xh) < 10:
            continue
        completed = sum(1 for c in video_cols if '已完成' in str(ws.cell(r, c).value or ''))
        if total_videos > 0:
            scores[xh] = round(completed / total_videos * 20, 1)
    return scores

def read_pdf_scores(pdf_path):
    """从PDF成绩登记表读平时/期末/总评 → {xh: {luru_ps, luru_qm, luru_zp}}"""
    scores = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table[2:]:
                    if not row or len(row) < 5:
                        continue
                    try:
                        xh = normalize_xh(row[1])
                        if not xh or len(xh) < 10:
                            continue
                        scores[xh] = {
                            'luru_ps': float(row[3]) if row[3] else 0,
                            'luru_qm': float(row[4]) if row[4] else 0,
                            'luru_zp': float(row[5]) if len(row) > 5 and row[5] else 0,
                        }
                    except (ValueError, TypeError, IndexError):
                        continue
    return scores

def read_names(path):
    """从综合成绩读姓名 → {xh: name}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    names = {}
    for r in range(4, ws.max_row + 1):
        xh = normalize_xh(ws.cell(r, 3).value)
        if not xh or len(xh) < 10:
            continue
        name = str(ws.cell(r, 2).value or '').strip()
        if name:
            names[xh] = name
    return names


def aggregate(course_dir):
    """归集所有源文件，输出统一的学生分数"""
    print(f"📂 扫描: {course_dir}")
    
    # 1. 平时成绩报告单 (互金251主源)
    pscj_path = os.path.join(course_dir, "互金251+工管251平时成绩报告单 .xlsx")
    pscj = read_pscj(pscj_path) if os.path.exists(pscj_path) else {}
    print(f"  平时成绩报告单: {len(pscj)}人")
    
    # 2. 班级和姓名信息
    zhcj_path = os.path.join(course_dir, "互金251+工管251_综合成绩.xlsx")
    names = {}
    classes = {}
    if os.path.exists(zhcj_path):
        wb = openpyxl.load_workbook(zhcj_path, data_only=True)
        ws = wb.active
        for r in range(4, ws.max_row + 1):
            xh = normalize_xh(ws.cell(r, 3).value)
            if not xh or len(xh) < 10:
                continue
            name = str(ws.cell(r, 2).value or '').strip()
            cls = str(ws.cell(r, 7).value or '').strip()
            if name: names[xh] = name
            if cls: classes[xh] = cls
    
    # 3. 手册成绩
    sc_path = os.path.join(course_dir, "互金251+工管251手册成绩.xlsx")
    shouce = read_shouce(sc_path) if os.path.exists(sc_path) else {}
    print(f"  手册成绩: {len(shouce)}人")
    
    # 4. 课堂表现
    kt_path = os.path.join(course_dir, "互金251+工管251_课堂表现成绩.xlsx")
    ketang = read_ketang(kt_path) if os.path.exists(kt_path) else {}
    print(f"  课堂表现: {len(ketang)}人")
    
    # 5. 阶段测试
    stage_dir = os.path.join(course_dir, "阶段测汇总试\\互金251工管251阶段测试成绩汇总\\互金251工管251阶段测试成绩汇总")
    stage = read_stage_tests(stage_dir) if os.path.exists(stage_dir) else {}
    print(f"  阶段测试: {len(stage)}人")
    
    # 6. 线上学习
    online_path = os.path.join(course_dir, "线上学习\\互金251+工管251线上成绩报告单.xlsx")
    online = read_online(online_path) if os.path.exists(online_path) else {}
    print(f"  线上学习: {len(online)}人")
    
    # 7. PDF成绩登记表
    pdf_path = r"C:\Users\Administrator\Desktop\大学计算机基础\大学计算机基础-教学文件-2025-2026(二）\课程档案\计算机2个班成绩登记表.pdf"
    pdf_data = read_pdf_scores(pdf_path) if os.path.exists(pdf_path) else {}
    print(f"  PDF成绩: {len(pdf_data)}人")
    
    # 汇总所有学生
    all_xh = set(list(pscj.keys()) + list(ketang.keys()) + list(shouce.keys()))
    students = []
    
    for xh in sorted(all_xh):
        if xh not in classes:
            continue
        
        # 阶段测试均从源文件读取独立4个成绩（阶段测试1→目标1, 测试2→目标2...）
        st_scores = stage.get(xh, {
            '阶段测试1': 0, '阶段测试2': 0, '阶段测试3': 0, '阶段测试4': 0
        })
        
        # 优先用平时成绩报告单
        if xh in pscj:
            s = pscj[xh].copy()
            # 移除折合后的阶段测试，替换为独立4个成绩
            del s['raw_scores']['阶段测试']
            s['raw_scores'].update(st_scores)
        else:
            # 从各子项源文件构建
            s = {
                'name': names.get(xh, xh),
                'raw_scores': {
                    '线上学习': online.get(xh, 0),
                    '课堂表现': ketang.get(xh, 0),
                    '学习手册': shouce.get(xh, 0),
                }
            }
            s['raw_scores'].update(st_scores)
        
        # 补充期末分数和录入成绩
        s['raw_scores']['期末考试'] = pdf_data.get(xh, {}).get('luru_qm', 0)
        s['luru_ps'] = pdf_data.get(xh, {}).get('luru_ps', 0)
        s['luru_qm'] = pdf_data.get(xh, {}).get('luru_qm', 0)
        s['luru_zp'] = pdf_data.get(xh, {}).get('luru_zp', 0)
        s['xh'] = xh
        s['class'] = classes.get(xh, '')
        
        students.append(s)
    
    print(f"\n📊 归集完成: {len(students)}名学生")
    return students


if __name__ == '__main__':
    course_dir = r"C:\Users\Administrator\Desktop\大学计算机基础\大学计算机基础-教学文件-2025-2026(二）\课程档案\互金251+工管251成绩 (1)\互金251+工管251成绩"
    students = aggregate(course_dir)
    
    # 快速对比
    print("\n=== 对比已有config ===")
    with open('skills/course_achievement_analysis/scripts/_temp/config_大学计算机基础.json', 'r', encoding='utf-8') as f:
        old_cfg = json.load(f)
    old_map = {s['xh']: s for s in old_cfg['students']}
    
    diffs = 0
    for s in students[:5]:
        xh = s['xh']
        new_rs = s['raw_scores']
        if xh in old_map:
            old_rs = old_map[xh]['raw_scores']
            for k in ['线上学习', '课堂表现',
                       '阶段测试1', '阶段测试2', '阶段测试3', '阶段测试4',
                       '学习手册', '期末考试']:
                nv = new_rs.get(k, 0)
                ov = old_rs.get(k, 0)
                if abs(nv - ov) > 0.5:
                    print(f"  {s['name']} {k}: old={ov} → new={nv}  🔴")
                    diffs += 1
    if diffs == 0:
        print(f"  ✅ 前5人分数一致")
    
    print(f"\n  互金251: {sum(1 for s in students if '互金' in s.get('class',''))}人")
    print(f"  工管251: {sum(1 for s in students if '工管' in s.get('class',''))}人")
    
    # 输出完整对比
    print(f"\n=== 所有差异 >0.5 ===")
    diffs = 0
    for s in students:
        xh = s['xh']
        new_rs = s['raw_scores']
        if xh in old_map:
            old_rs = old_map[xh]['raw_scores']
            has_diff = False
            for k in ['线上学习', '课堂表现',
                       '阶段测试1', '阶段测试2', '阶段测试3', '阶段测试4',
                       '学习手册', '期末考试']:
                nv = new_rs.get(k, 0)
                ov = old_rs.get(k, 0)
                if abs(nv - ov) > 0.5:
                    if not has_diff:
                        print(f"\n  {s['name']}({xh}, {s['class']}):")
                        has_diff = True
                    print(f"    {k}: {ov} → {nv}")
                    diffs += 1
    print(f"\n  总计差异项: {diffs}")
