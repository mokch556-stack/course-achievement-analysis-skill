#!/usr/bin/env python3
"""
verify_xlsx.py — 达成度计算表三核查验证脚本
功能：自动检查xlsx的结构完整性和数据一致性
输出：PASS/FAIL 报告，明确写出每项检查结果

操作时间: 2026-06-28
操作agent: default
"""

import sys
import json
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ 需要 openpyxl 库：pip install openpyxl")
    sys.exit(1)


def load_xlsx(path):
    """加载xlsx文件，返回 workbook"""
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb


def check_sheet_structure(ws):
    """检查主工作表的结构完整性：毕业要求行/教学目标行/考核项行/满分值行/平均行/达成度行/分班行"""
    errors = []
    max_row = ws.max_row
    max_col = ws.max_column

    # 检查行数
    if max_row < 12:
        errors.append(f"❌ 行数过少（{max_row}行），缺少必要的汇总行")

    # 收集所有行的文本内容
    row_texts = {}
    for r in range(1, max_row+1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col+1)]
        row_texts[r] = ' '.join(str(v or '') for v in row_vals)

    # 检查毕业要求行
    found_indicator = any('毕业要求' in row_texts[r] for r in range(1, 15))
    if not found_indicator:
        errors.append("❌ 毕业要求指标点行缺失或内容不完整")

    # 检查教学目标行
    found_target = any('教学目标' in row_texts[r] for r in range(1, 15))
    if not found_target:
        errors.append("❌ 课程教学目标行缺失或内容不完整")

    # 检查折合满分行
    found_full = any('折合满分' in row_texts[r] for r in range(1, 10))
    if not found_full:
        errors.append("❌ 折合满分信息行缺失")

    # 检查列数（应该是14列）
    if max_col != 14:
        errors.append(f"❌ 列数应为14（当前{max_col}列），平时子项应单独成列，期末不分子项")

    # 检查最后几行是否有平均行
    last_rows = [row_texts.get(r, '') for r in range(max_row-5, max_row+1)]
    has_avg = any('平均' in row for row in last_rows)
    has_achieve = any('达成度' in row for row in last_rows)
    has_class = any('达成度' in row for row in last_rows)

    if not has_avg:
        errors.append("❌ 缺少平均分行（未找到含'平均'的行）")
    if not has_achieve:
        errors.append("❌ 缺少目标和/或课程达成度行（未找到含'达成度'的行）")

    return errors


def check_data_consistency(ws, chengji_path=None):
    """检查数据一致性：折合分、学生数量等"""
    errors = []
    max_row = ws.max_row
    max_col = ws.max_column

    # 找折合分列（最后一列）
    zhehe_col = max_col  # 折合分通常在最右列

    # 找数据起始行（找第一行有数字序号的行）
    data_start = None
    for r in range(5, max_row+1):
        val = ws.cell(r, 1).value
        if isinstance(val, (int, float)) and val > 0 and val < 200:
            data_start = r
            break
    if data_start is None:
        errors.append("❌ 未找到数据起始行")
        return errors

    # 检查折合分是否合理（0-100之间）
    zhehe_values = []
    for r in range(data_start, max_row+1):
        val = ws.cell(r, zhehe_col).value
        if isinstance(val, (int, float)):
            zhehe_values.append(val)

    if zhehe_values:
        # 检查是否所有值都在0-100之间
        out_of_range = [v for v in zhehe_values if v < 0 or v > 100]
        if out_of_range:
            errors.append(f"❌ 有{len(out_of_range)}个学生的折合分超出0-100范围")
        # 检查是否有负值
        negatives = [v for v in zhehe_values if v < 0]
        if negatives:
            errors.append(f"❌ 有{len(negatives)}个学生的折合分为负值")
        # 检查是否有None
        none_count = sum(1 for v in zhehe_values if v is None)
        if none_count > 0:
            errors.append(f"❌ 有{none_count}个学生的折合分为空")
    else:
        errors.append("❌ 未能读取到任何折合分数据")

    return errors


def check_score_alignment(ws, luru_path=None):
    """
    检查折合分与录入成绩总评是否对齐
    如果提供了录入成绩文件路径，进行对比验证
    """
    if not luru_path:
        return ["⚠️ 未提供录入成绩文件路径，跳过折合分对齐检查"]
    
    errors = []
    try:
        luru_wb = openpyxl.load_workbook(luru_path, data_only=True)
        luru_ws = luru_wb.active
        
        # 获取学生名单和总评成绩
        luru_data = {}
        for row in luru_ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[2]:  # 学号 + 姓名
                xuehao = str(row[0]).strip()
                name = str(row[2]).strip()
                zongping = row[6] if len(row) > 6 else None  # col 7 = 总评
                if zongping is not None:
                    luru_data[name] = float(zongping) if isinstance(zongping, (int, float)) else None
        
        # 对比xlsx中的折合分
        max_row = ws.max_row
        zhehe_col = ws.max_column
        
        mismatches = []
        for r in range(8, max_row + 1):
            name = ws.cell(r, 3).value  # col 3 = 姓名
            zhehe = ws.cell(r, zhehe_col).value
            
            if name and isinstance(zhehe, (int, float)):
                name = str(name).strip()
                if name in luru_data and luru_data[name] is not None:
                    diff = abs(zhehe - luru_data[name])
                    if diff > 0.5:
                        mismatches.append((name, zhehe, luru_data[name], diff))
        
        if mismatches:
            errors.append(f"❌ 有{len(mismatches)}名学生折合分与录入总评不一致（误差>0.5）：")
            for name, zhehe, luru, diff in mismatches[:5]:
                errors.append(f"    {name}: 折合{zhehe} ≠ 录入{luru} (差{diff})")
            if len(mismatches) > 5:
                errors.append(f"    ...及其他{len(mismatches)-5}人")
        else:
            errors.append("✅ 所有学生折合分与录入总评一致 ✓")
            
    except Exception as e:
        errors.append(f"⚠️ 读取录入成绩文件失败: {e}")
    
    return errors


def main():
    parser = argparse.ArgumentParser(description='达成度计算表验证脚本')
    parser.add_argument('xlsx_path', help='达成度计算表.xlsx 路径')
    parser.add_argument('--luru', help='录入成绩.xlsx 路径（用于折合分对齐检查）')
    parser.add_argument('--json', action='store_true', help='以JSON格式输出结果')
    args = parser.parse_args()

    if not Path(args.xlsx_path).exists():
        print(f"❌ 文件不存在: {args.xlsx_path}")
        sys.exit(1)

    wb = load_xlsx(args.xlsx_path)

    # 查找主工作表（Sheet1或达成度计算表）
    main_sheet = None
    for sn in ['Sheet1', '达成度计算表', '达成度计算']:
        if sn in wb.sheetnames:
            main_sheet = sn
            break
    if main_sheet is None:
        result = {"status": "FAIL", "errors": ["❌ 缺少主工作表(Sheet1/达成度计算表)"], "warnings": []}
        if args.json:
            print(json.dumps(result, ensure_ascii=False, indent=2))
        else:
            print("=" * 50)
            print("📋 达成度计算表验证报告")
            print("=" * 50)
            print(f"\n❌ 严重错误：找不到主工作表（可用: {wb.sheetnames}）")
        sys.exit(1)

    ws = wb[main_sheet]
    
    # 执行检查
    struct_errors = check_sheet_structure(ws)
    data_errors = check_data_consistency(ws)
    align_errors = check_score_alignment(ws, Path(args.luru) if args.luru else None)

    all_errors = struct_errors + data_errors + align_errors
    has_error = any(e.startswith('❌') for e in all_errors)
    has_warning = any(e.startswith('⚠️') for e in all_errors)

    result = {
        "status": "FAIL" if has_error else ("WARN" if has_warning else "PASS"),
        "errors": [e for e in all_errors if e.startswith('❌')],
        "warnings": [e for e in all_errors if e.startswith('⚠️')],
        "pass": [e for e in all_errors if e.startswith('✅')],
        "summary": {
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "data_rows": ws.max_row - 7,
            "total_errors": len(all_errors)
        }
    }

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print("=" * 50)
        print("📋 达成度计算表验证报告")
        print("=" * 50)
        print(f"\n📊 基本信息")
        print(f"  · 工作表数: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")
        print(f"  · 数据行数: {ws.max_row - 7} (学生)")
        print(f"  · 表列数:   {ws.max_column}")
        print(f"\n📝 结构检查:")
        for e in struct_errors:
            print(f"  {e}")
        print(f"\n📊 数据检查:")
        for e in data_errors:
            print(f"  {e}")
        print(f"\n📐 对齐检查:")
        for e in align_errors:
            print(f"  {e}")
        print(f"\n{'=' * 50}")
        if has_error:
            print(f"🔴 总体状态: FAIL（{len([e for e in all_errors if e.startswith('❌')])}个错误）")
        elif has_warning:
            print(f"🟡 总体状态: WARN（有警告，需人工确认）")
        else:
            print(f"🟢 总体状态: PASS（全部通过 ✓）")
        print(f"{'=' * 50}")


if __name__ == '__main__':
    main()
