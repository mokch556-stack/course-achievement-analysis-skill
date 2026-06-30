"""
probe_course_dir.py — 探查课程档案目录，列出所有源文件及其结构
> 操作时间: 2026-06-29
> 操作agent: default

输出：每个文件的 Sheet 名、行数、列名、第一行数据
"""
import openpyxl, pdfplumber, os, sys

def probe_xlsx(path, label=""):
    """探查 xlsx 文件"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    print(f"\n📄 {label}")
    print(f"   路径: {path}")
    print(f"   Sheet: {wb.sheetnames}")
    print(f"   行列: {ws.max_row}行 × {ws.max_column}列")
    
    # 找表头行（有连续非空值的行）
    for r in range(1, min(10, ws.max_row+1)):
        headers = []
        for c in range(1, min(30, ws.max_column+1)):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                headers.append((c, str(v).strip()[:30]))
        if len(headers) >= 3:
            print(f"   表头(R{r}):")
            for c, h in headers:
                print(f"     C{c}: {h}")
            # 下一行数据
            if r+1 <= ws.max_row:
                data = []
                for c in range(1, min(30, ws.max_column+1)):
                    v = ws.cell(r+1, c).value
                    if v is not None and str(v).strip():
                        data.append((c, str(v).strip()[:30]))
                if data:
                    print(f"   数据(R{r+1}):")
                    for c, d in data[:8]:
                        print(f"     C{c}: {d}")
            break
    wb.close()


def probe_pdf(path, label=""):
    """探查 PDF 文件"""
    print(f"\n📄 {label}")
    print(f"   路径: {path}")
    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            if tables:
                print(f"   第{pi+1}页: {len(tables)}个表格")
                for ti, t in enumerate(tables):
                    if t:
                        print(f"     表格{ti+1}: {len(t)}行×{len(t[0])}列")
                        if len(t) >= 2:
                            print(f"       表头: {t[1][:6]}")
                            if len(t) >= 3:
                                print(f"       数据: {t[3][:6]}")


def main(course_dir):
    print("=" * 60)
    print("📂 课程档案目录探查")
    print(f"   {course_dir}")
    print("=" * 60)
    
    for root, dirs, files in os.walk(course_dir):
        # 跳过好中差子目录
        dirs[:] = [d for d in dirs if d not in ['好', '中', '差']]
        
        dirs.sort()
        for f in sorted(files):
            path = os.path.join(root, f)
            rel = os.path.relpath(path, course_dir)
            if f.endswith('.xlsx') and not f.startswith('~$'):
                probe_xlsx(path, rel)
            elif f.endswith('.pdf'):
                probe_pdf(path, rel)


if __name__ == '__main__':
    course_dir = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\Administrator\Desktop\大学计算机基础\大学计算机基础-教学文件-2025-2026(二）\课程档案\互金251+工管251成绩 (1)\互金251+工管251成绩"
    main(course_dir)
